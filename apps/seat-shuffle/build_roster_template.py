"""席替えアプリ 名簿テンプレート（.xlsx）生成スクリプト

アプリの名簿取り込みが期待する列（出席番号・氏名・ふりがな・性別）に合わせた
空のテンプレートを作成する。記入後、アプリの「Excelファイルを読み込む」で取り込める。
アプリ内の「名簿テンプレート（Excel）をダウンロード」ボタンと同じ内容。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = "名簿テンプレート.xlsx"
HEADERS = ["出席番号", "氏名", "ふりがな", "性別"]
EXAMPLES = [
    [1, "山田 太郎", "やまだ たろう", "男"],
    [2, "佐藤 花子", "さとう はなこ", "女"],
]
LAST_NUMBER = 40  # 出席番号をあらかじめ入れておく行数

PRIMARY = "2D6A4F"
LIGHT = "E8F4EE"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "名簿"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ヘッダー行
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.alignment = center
        cell.border = border

    # 記入例＋出席番号入りの空行
    rows = list(EXAMPLES)
    for n in range(len(EXAMPLES) + 1, LAST_NUMBER + 1):
        rows.append([n, "", "", ""])

    for r_off, row in enumerate(rows):
        excel_row = r_off + 2
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c, value=val)
            cell.border = border
            cell.alignment = center if c in (1, 4) else left
        # 記入例の行はうすい色で目立たせる
        if r_off < len(EXAMPLES):
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=excel_row, column=c).fill = PatternFill("solid", fgColor=LIGHT)

    # 列幅・行高・ヘッダー固定
    for col, width in zip("ABCD", (10, 20, 22, 8)):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # 使い方メモ（取り込みに影響しない位置・ヘッダー誤認しない文言）
    note = ws.cell(
        row=1, column=6,
        value="この行が見出しです。氏名は必須、ふりがな・性別は任意（性別は 男／女）。"
              "記入例の2行は上書きまたは削除して使ってください。",
    )
    note.font = Font(color="888888", size=9)
    ws.column_dimensions["F"].width = 60

    wb.save(OUT)
    print(f"Generated: {OUT}")


if __name__ == "__main__":
    main()
