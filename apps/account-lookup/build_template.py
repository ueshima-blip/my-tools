"""アカウント確認ツール 名簿テンプレート（.xlsx）生成スクリプト

アプリ（index.html）が取り込む列に合わせた空のテンプレートを作成する。
記入後、アプリの「① Excel名簿を読み込む」で取り込める。
アプリ内の「テンプレート（Excel）をダウンロード」ボタンと同じ内容。

レイアウト:
    1行目 A1 : 「まなびポケット学校コード（全員共通）」ラベル
    2行目 A2 : 学校コードの値（全員共通の1つだけ入力）
    3行目    : 見出し（年 / 組 / 番 / 氏名（漢字） / 氏名（ひらがな）
               / Googleアカウント / Googleパスワード / 端末番号 / 端末暗証番号）
    4行目〜  : 生徒データ（1人1行）

再生成: このフォルダで `python3 build_template.py`
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = "アカウント名簿テンプレート.xlsx"

HEADERS = [
    "年", "組", "番", "氏名（漢字）", "氏名（ひらがな）",
    "Googleアカウント", "Googleパスワード", "端末番号", "端末暗証番号",
]

# 全員共通のまなびポケット学校コード（架空の記入例）
SCHOOL_CODE_LABEL = "まなびポケット学校コード（全員共通）"
SCHOOL_CODE_SAMPLE = "manabi-1234"

# 記入例（すべて架空。使うときは上書き／削除してください）
EXAMPLES = [
    [1, 1, 1, "山田 太郎", "やまだ たろう",
     "taro.yamada@example-jhs.ed.jp", "Yamada-1234", "T-001", "0042"],
    [1, 1, 2, "佐藤 花子", "さとう はなこ",
     "hanako.sato@example-jhs.ed.jp", "Sato-5678", "T-002", "0157"],
    [2, 3, 15, "鈴木 一郎", "すずき いちろう",
     "ichiro.suzuki@example-jhs.ed.jp", "Suzuki-9012", "T-115", "0830"],
]

EMPTY_ROWS = 40  # 記入例のあとに用意しておく空行の数

HEADER_ROW = 3   # 見出しの行
# 文字列として扱う列（先頭の 0 や記号を保つため）: 1始まりの列番号
TEXT_COLS = {6, 7, 8, 9}  # アカウント / パスワード / 端末番号 / 端末暗証番号

PRIMARY = "1F4E79"   # 濃い青（見出し）
LIGHT = "E7EFF7"     # うすい青（記入例の目印）
ACCENT = "C8531B"    # オレンジ（学校コード）
ACCENT_BG = "FCEEE3"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "名簿"

    thin = Side(style="thin", color="BFC9D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # --- 学校コード（全員共通）: A1 ラベル / A2 値 ---
    c1 = ws.cell(row=1, column=1, value=SCHOOL_CODE_LABEL)
    c1.font = Font(bold=True, color=ACCENT, size=11)
    c2 = ws.cell(row=2, column=1)
    c2.number_format = "@"            # 文字列
    c2.value = SCHOOL_CODE_SAMPLE
    c2.font = Font(bold=True, size=11)
    c2.fill = PatternFill("solid", fgColor=ACCENT_BG)
    c2.alignment = left
    c2.border = border

    # --- 見出し行（3行目） ---
    for c, head in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c, value=head)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.alignment = center
        cell.border = border

    # --- 記入例 ＋ 空行（4行目〜） ---
    rows = list(EXAMPLES) + [[""] * len(HEADERS) for _ in range(EMPTY_ROWS)]
    for r_off, row in enumerate(rows):
        excel_row = HEADER_ROW + 1 + r_off
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c)
            if c in TEXT_COLS:
                cell.number_format = "@"  # 文字列（0042 などの先頭ゼロを保つ）
                cell.value = "" if val == "" else str(val)
            else:
                cell.value = val
            cell.border = border
            cell.alignment = center if c in (1, 2, 3) else left
        if r_off < len(EXAMPLES):  # 記入例はうすい色で目立たせる
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=excel_row, column=c).fill = PatternFill("solid", fgColor=LIGHT)

    # 列幅
    widths = {
        "A": 6, "B": 6, "C": 6, "D": 16, "E": 18,
        "F": 30, "G": 16, "H": 12, "I": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[HEADER_ROW].height = 24
    ws.freeze_panes = f"A{HEADER_ROW + 1}"  # 学校コード＋見出しを固定

    # 使い方メモ（取り込みに影響しない右側の列に置く）
    note_col = len(HEADERS) + 2  # K列
    note1 = ws.cell(
        row=1, column=note_col,
        value="◀ まなびポケット学校コードは「全員共通」です。すぐ下のセル（A2）に1つだけ入力してください。",
    )
    note1.font = Font(color="808080", size=9)
    note1.alignment = Alignment(vertical="center")
    note3 = ws.cell(
        row=HEADER_ROW, column=note_col,
        value="◀ この行が見出しです（消さないでください）。「氏名（ひらがな）」で検索します。"
              "記入例の3行は上書きまたは削除して使ってください。"
              "端末暗証番号などの先頭の 0 は、この列が『文字列』設定なのでそのまま保てます。",
    )
    note3.font = Font(color="808080", size=9)
    note3.alignment = Alignment(vertical="center")
    ws.column_dimensions[chr(ord("A") + note_col - 1)].width = 76

    wb.save(OUT)
    print(f"Generated: {OUT}  (学校コード行 + {len(HEADERS)} columns, "
          f"{len(EXAMPLES)} examples + {EMPTY_ROWS} blank rows)")


if __name__ == "__main__":
    main()
